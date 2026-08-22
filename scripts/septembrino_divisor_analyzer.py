#!/usr/bin/env python3
"""
septembrino_divisor_analyzer.py
Единый анализатор матричных данных Septembrino.
Извлекает высокие делители, фильтрует по классам вычетов, 
преобразует в степени двойки и генерирует CSV/Excel для отправки Anabel.

Требования:
    pip install openpyxl

Запуск:
    python septembrino_divisor_analyzer.py
"""

import os
import csv
import math
import sys
from collections import defaultdict
from typing import Dict, List, Tuple, Callable

# Проверка зависимости
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl не установлен. Excel-отчёт будет пропущен.")
    print("   Установите: pip install openpyxl")

# ============================================================================
# КОНФИГУРАЦИЯ
# ============================================================================
INPUT_FILE = 'septembrino_table.csv'
OUTPUT_DIR = 'output'
MAX_M = 60
MIN_DIVISOR = 32
REMOVE_DIVISORS = {2, 4, 8, 16}

# Классы вычетов для анализа
RESIDUE_CLASSES: Dict[str, Callable[[int], bool]] = {
    'k_eq_1_mod_32': lambda k: k % 32 == 1,
    'k_eq_17_mod_32': lambda k: k % 32 == 17,
    'k_eq_3_mod_16': lambda k: k % 16 == 3,
}

# Цветовая карта для степеней (log2)
COLOR_MAP = {
    5: '92D050',   # 32
    6: 'C6E0B4',   # 64
    7: 'FFEB9C',   # 128
    8: 'FFC000',   # 256
    9: 'FF9900',   # 512
    10: 'FF6600',  # 1024
    11: 'FF3300',  # 2048
    12: 'FF0000',  # 4096
    13: 'CC0000',  # 8192
    14: '990000',  # 16384
    15: '660000',  # 32768
    16: '330000',  # 65536
    17: '000000',  # 131072+
}

# ============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================================
def log2_safe(d: int) -> int:
    """Безопасное вычисление log2 для степеней двойки."""
    if d <= 0:
        return 0
    return int(math.log2(d))

def get_color(exp: int) -> str:
    """Возвращает HEX-цвет для степени."""
    if exp in COLOR_MAP:
        return COLOR_MAP[exp]
    return COLOR_MAP[17]  # Для >65536

def ensure_dir(path: str):
    """Создаёт директорию, если не существует."""
    os.makedirs(path, exist_ok=True)

# ============================================================================
# ЗАГРУЗКА И ПАРСИНГ ДАННЫХ
# ============================================================================
def load_trajectories(filepath: str) -> List[Dict]:
    """Загружает septembrino_table.csv и возвращает список словарей."""
    if not os.path.exists(filepath):
        print(f"❌ Файл {filepath} не найден.")
        sys.exit(1)
        
    data = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                k = int(row['k'])
                m = int(row['m'])
                divs = []
                for col in sorted(row.keys()):
                    if col.startswith('div_') and row[col]:
                        try:
                            d = int(row[col])
                            if d >= MIN_DIVISOR and d not in REMOVE_DIVISORS:
                                divs.append(d)
                        except ValueError:
                            continue
                data.append({'k': k, 'm': m, 'divisors': divs})
            except (KeyError, ValueError):
                continue
    print(f"✅ Загружено {len(data)} записей.")
    return data

# ============================================================================
# ГРУППИРОВКА ПО КЛАССАМ
# ============================================================================
def group_by_class(data: List[Dict]) -> Dict[str, List[Dict]]:
    """Группирует данные по классам вычетов."""
    groups = {name: [] for name in RESIDUE_CLASSES}
    for row in data:
        k = row['k']
        for name, predicate in RESIDUE_CLASSES.items():
            if predicate(k):
                groups[name].append(row)
                break  # Каждое k попадает только в один класс
    return groups

# ============================================================================
# ГЕНЕРАЦИЯ CSV
# ============================================================================
def save_csv(groups: Dict[str, List[Dict]], out_dir: str):
    """Сохраняет CSV файлы с степенями вместо значений."""
    ensure_dir(out_dir)
    for name, rows in groups.items():
        if not rows:
            continue
            
        filepath = os.path.join(out_dir, f'{name}.csv')
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # Заголовок: k, m0, m1, ..., m60
            header = ['k'] + [f'm{i}' for i in range(MAX_M + 1)]
            writer.writerow(header)
            
            # Сортировка по k
            rows.sort(key=lambda x: x['k'])
            
            for row in rows:
                k = row['k']
                m = row['m']
                divs = row['divisors']
                
                # Создаём строку: k, затем степени для каждого m
                row_data = [k] + ['.'] * (MAX_M + 1)
                for d in divs:
                    exp = log2_safe(d)
                    if m <= MAX_M:
                        row_data[m + 1] = str(exp)
                writer.writerow(row_data)
        print(f"📄 Сохранён CSV: {filepath}")

# ============================================================================
# ГЕНЕРАЦИЯ EXCEL
# ============================================================================
def save_excel(groups: Dict[str, List[Dict]], out_dir: str):
    """Создаёт Excel-книгу с форматированными таблицами."""
    if not EXCEL_AVAILABLE:
        return
        
    ensure_dir(out_dir)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Удаляем пустой лист
    
    for name, rows in groups.items():
        if not rows:
            continue
            
        ws = wb.create_sheet(title=name.replace('k_eq_', '').replace('_mod_', ' mod '))
        
        # Заголовки
        ws['A1'] = f'Divisor Exponents: k ≡ {name.split("_mod_")[1].replace("_", " ")}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f'Removing: 2, 4, 8, 16 | Showing: ≥ {MIN_DIVISOR} (as log2)'
        ws['A2'].font = Font(italic=True)
        
        # Столбцы m=0..60
        ws.cell(row=4, column=1).value = 'k'
        for m in range(MAX_M + 1):
            cell = ws.cell(row=4, column=m + 2)
            cell.value = m
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
        # Данные
        rows.sort(key=lambda x: x['k'])
        for idx, row in enumerate(rows):
            r = idx + 5
            k = row['k']
            m = row['m']
            divs = row['divisors']
            
            ws.cell(row=r, column=1).value = k
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')
            
            # Заполняем m-столбцы
            for col_idx in range(2, MAX_M + 3):
                cell = ws.cell(row=r, column=col_idx)
                cell.value = '.'
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(size=9, color='999999')
                
            for d in divs:
                if m <= MAX_M:
                    exp = log2_safe(d)
                    col = m + 2
                    cell = ws.cell(row=r, column=col)
                    cell.value = exp
                    
                    # Цвет и шрифт
                    color = get_color(exp)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    if exp >= 12:
                        cell.font = Font(bold=True, color='FFFFFF', size=9)
                    else:
                        cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal='center')
                    
            ws.row_dimensions[r].height = 15
            
        # Ширина столбцов
        ws.column_dimensions['A'].width = 8
        for m in range(MAX_M + 1):
            col_letter = get_column_letter(m + 2)
            ws.column_dimensions[col_letter].width = 5
            
        ws.freeze_panes = 'B5'
        print(f"📊 Лист Excel: {name}")
        
    # Легенда
    ws_legend = wb.create_sheet('Legend')
    ws_legend['A1'] = 'Exponent → Divisor → Color'
    ws_legend['A1'].font = Font(bold=True)
    r = 3
    for exp, color in sorted(COLOR_MAP.items()):
        d = 2 ** exp
        ws_legend.cell(row=r, column=1).value = exp
        ws_legend.cell(row=r, column=2).value = d
        cell = ws_legend.cell(row=r, column=3)
        cell.value = '■'
        cell.font = Font(color=color, size=14)
        r += 1
        
    filepath = os.path.join(out_dir, 'septembrino_divisors.xlsx')
    wb.save(filepath)
    print(f"💾 Сохранён Excel: {filepath}")

# ============================================================================
# СВОДНАЯ СТАТИСТИКА
# ============================================================================
def print_summary(groups: Dict[str, List[Dict]]):
    """Выводит статистику в консоль."""
    print("\n" + "="*60)
    print("📈 СВОДНАЯ СТАТИСТИКА")
    print("="*60)
    print(f"{'Класс':<20} | {'k':<5} | {'Записей':<8} | {'Макс делитель':<12} | {'Средний':<8}")
    print("-"*60)
    
    for name, rows in groups.items():
        if not rows:
            continue
        k_count = len(set(r['k'] for r in rows))
        total = len(rows)
        all_divs = [d for r in rows for d in r['divisors']]
        max_d = max(all_divs) if all_divs else 0
        avg_d = sum(all_divs)/len(all_divs) if all_divs else 0
        print(f"{name:<20} | {k_count:<5} | {total:<8} | {max_d:<12} | {avg_d:.1f}")
    print("="*60)

# ============================================================================
# ГЛАВНЫЙ ЗАПУСК
# ============================================================================
def main():
    print("🚀 Запуск анализатора матриц Septembrino...")
    print(f"📂 Вход: {INPUT_FILE}")
    print(f"📁 Выход: {OUTPUT_DIR}/")
    print("-"*40)
    
    # 1. Загрузка
    data = load_trajectories(INPUT_FILE)
    
    # 2. Группировка
    groups = group_by_class(data)
    
    # 3. CSV
    save_csv(groups, OUTPUT_DIR)
    
    # 4. Excel
    save_excel(groups, OUTPUT_DIR)
    
    # 5. Статистика
    print_summary(groups)
    
    print("\n✅ Готово! Файлы сохранены в папке 'output/'.")
    print("📤 Отправляйте Anabel: septembrino_divisors.xlsx + CSV файлы.")

if __name__ == '__main__':
    main()