#!/usr/bin/env python3
"""
Create Excel tables from Septembrino filtered divisor data.
Generates clean, readable spreadsheets with conditional formatting.

Usage:
    python create_excel_tables.py

Requirements:
    pip install openpyxl
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================================
# CONFIGURATION
# ============================================================================

INPUT_FILES = [
    ('septembrino_filtered_k_eq_3_mod_16.txt', 'k ≡ 3 (mod 16)'),
    ('septembrino_filtered_k_eq_1_mod_32.txt', 'k ≡ 1 (mod 32)'),
    ('septembrino_filtered_k_eq_17_mod_32.txt', 'k ≡ 17 (mod 32)'),
]

OUTPUT_FILE = 'septembrino_divisor_tables.xlsx'
MAX_M = 60
MIN_DIVISOR = 32

# Color scale for divisors (log scale)
DIVISOR_COLORS = {
    32: '92D050',      # Light green
    64: 'C6E0B4',      # Green
    128: 'FFEB9C',     # Light yellow
    256: 'FFC000',     # Yellow-orange
    512: 'FF9900',     # Orange
    1024: 'FF6600',    # Dark orange
    2048: 'FF3300',    # Red-orange
    4096: 'FF0000',    # Red
    8192: 'CC0000',    # Dark red
    16384: '990000',   # Very dark red
    32768: '660000',   # Deep red
    65536: '330000',   # Darkest red
}

# ============================================================================
# Parse filtered files
# ============================================================================

def parse_filtered_file(filename):
    """
    Parse the vertical list format.
    Returns: dict[k] = {m: highest_divisor}
    """
    data = {}
    
    if not os.path.exists(filename):
        print(f"WARNING: {filename} not found!")
        return data
    
    k_m_counter = {}
    
    with open(filename, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            
            if not line or line.startswith('DIVISORS') or line.startswith('=') or line.startswith('Range'):
                continue
            
            if ':' not in line:
                continue
            
            parts = line.split(':', 1)
            if len(parts) != 2:
                continue
            
            k_str = parts[0].strip()
            divisors_str = parts[1].strip()
            
            try:
                k = int(k_str)
            except ValueError:
                continue
            
            # Track m for this k
            if k not in k_m_counter:
                k_m_counter[k] = 0
            m = k_m_counter[k]
            k_m_counter[k] += 1
            
            # Parse divisors and take highest
            if divisors_str:
                divisors = divisors_str.split()
                highest = 0
                for d in divisors:
                    try:
                        divisor = int(d)
                        if divisor >= MIN_DIVISOR and divisor > highest:
                            highest = divisor
                    except ValueError:
                        continue
                
                if k not in data:
                    data[k] = {}
                if highest > 0:
                    data[k][m] = highest
    
    return data

# ============================================================================
# Create Excel workbook
# ============================================================================

def create_excel_workbook(all_data, output_file):
    """Create Excel workbook with formatted tables."""
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create summary sheet
    summary_ws = wb.create_sheet('Summary')
    _create_summary_sheet(summary_ws, all_data)
    
    # Create sheet for each residue class
    for filename, title in INPUT_FILES:
        data = all_data.get(filename, {})
        if data:
            ws = wb.create_sheet(title.replace(' ', '_').replace('≡', 'eq').replace('(', '').replace(')', ''))
            _create_divisor_table(ws, data, title)
    
    # Save workbook
    wb.save(output_file)
    print(f"Created {output_file}")

def _create_summary_sheet(ws, all_data):
    """Create summary statistics sheet."""
    
    # Header
    ws['A1'] = 'SEPTembrino DIVISOR ANALYSIS - SUMMARY'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = 'Residue Class'
    ws['B3'] = 'Total k Values'
    ws['C3'] = 'Total Entries'
    ws['D3'] = 'Max Divisor'
    ws['E3'] = 'Avg Divisor'
    ws['F3'] = 'High Divisors (≥4096)'
    
    # Style header
    for col in range(1, 7):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    row = 4
    for filename, title in INPUT_FILES:
        data = all_data.get(filename, {})
        if data:
            total_k = len(data)
            total_entries = sum(len(data[k]) for k in data)
            all_divs = [d for k in data for d in data[k].values()]
            max_div = max(all_divs) if all_divs else 0
            avg_div = sum(all_divs) / len(all_divs) if all_divs else 0
            high_divs = sum(1 for d in all_divs if d >= 4096)
            
            ws[f'A{row}'] = title
            ws[f'B{row}'] = total_k
            ws[f'C{row}'] = total_entries
            ws[f'D{row}'] = max_div
            ws[f'E{row}'] = f'{avg_div:.1f}'
            ws[f'F{row}'] = high_divs
            
            row += 1
    
    # Add notes
    row += 2
    ws[f'A{row}'] = 'Notes:'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = '- Data generated from k=1 to 2000, m=0 to 60'
    row += 1
    ws[f'A{row}'] = '- Only divisors ≥ 32 shown (2, 4, 8, 16 removed)'
    row += 1
    ws[f'A{row}'] = '- Each cell shows highest divisor for that (k, m) combination'
    row += 1
    ws[f'A{row}'] = '- Color scale: green (32) → red (65536+)'

def _create_divisor_table(ws, data, title):
    """Create divisor table with conditional formatting."""
    
    # Title
    ws['A1'] = f'Divisor Table: {title}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Removing: 2, 4, 8, 16 | Showing: ≥ {MIN_DIVISOR}'
    ws['A2'].font = Font(italic=True)
    
    # Header row (m values)
    ws['B4'] = 'k'
    for m in range(MAX_M + 1):
        col = m + 2  # B=2, C=3, ...
        ws.cell(row=4, column=col).value = m
        ws.cell(row=4, column=col).font = Font(bold=True, size=9)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=4, column=col).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Data rows
    sorted_k = sorted(data.keys())
    
    for row_idx, k in enumerate(sorted_k):
        excel_row = row_idx + 5  # Start at row 5
        k_data = data[k]
        
        # k value
        ws.cell(row=excel_row, column=1).value = k
        ws.cell(row=excel_row, column=1).font = Font(bold=True)
        ws.cell(row=excel_row, column=1).alignment = Alignment(horizontal='right')
        
        # Divisor values
        for m in range(MAX_M + 1):
            excel_col = m + 2
            cell = ws.cell(row=excel_row, column=excel_col)
            
            if m in k_data:
                d = k_data[m]
                cell.value = d
                
                # Apply color based on divisor value
                color = _get_divisor_color(d)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                # Font color for readability
                if d >= 4096:
                    cell.font = Font(bold=True, color='FFFFFF')
                else:
                    cell.font = Font(size=9)
                
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.value = '.'
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(size=9, color='999999')
        
        # Set row height
        ws.row_dimensions[excel_row].height = 15
    
    # Set column widths
    ws.column_dimensions['A'].width = 8  # k values
    for m in range(MAX_M + 1):
        col_letter = get_column_letter(m + 2)
        ws.column_dimensions[col_letter].width = 5  # m values
    
    # Freeze header row and k column
    ws.freeze_panes = 'B5'
    
    # Add legend
    legend_row = len(sorted_k) + 8
    ws[f'A{legend_row}'] = 'Legend:'
    ws[f'A{legend_row}'].font = Font(bold=True)
    
    legend_cols = [
        (32, '92D050'),
        (128, 'FFEB9C'),
        (512, 'FF9900'),
        (2048, 'FF3300'),
        (8192, 'CC0000'),
        (32768, '660000'),
    ]
    
    for i, (div, color) in enumerate(legend_cols):
        col = i + 1
        cell = ws.cell(row=legend_row + 1, column=col)
        cell.value = div
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF' if div >= 2048 else '000000')
        cell.alignment = Alignment(horizontal='center')

def _get_divisor_color(divisor):
    """Get color for divisor value."""
    thresholds = sorted(DIVISOR_COLORS.keys(), reverse=True)
    
    for threshold in thresholds:
        if divisor >= threshold:
            return DIVISOR_COLORS[threshold]
    
    return DIVISOR_COLORS[32]  # Default

# ============================================================================
# Main
# ============================================================================

def main():
    print("=" * 80)
    print("CREATING EXCEL TABLES FOR SEPTembrino DIVISOR DATA")
    print("=" * 80)
    print()
    
    # Parse all data
    all_data = {}
    for filename, title in INPUT_FILES:
        print(f"Loading {filename}...")
        data = parse_filtered_file(filename)
        all_data[filename] = data
        print(f"  Loaded {len(data)} k values")
    print()
    
    # Create Excel workbook
    print(f"Creating Excel workbook: {OUTPUT_FILE}...")
    create_excel_workbook(all_data, OUTPUT_FILE)
    print()
    
    print("=" * 80)
    print("DONE!")
    print("=" * 80)
    print()
    print(f"Open {OUTPUT_FILE} in Excel to view formatted tables.")
    print()
    print("Sheets:")
    print("  - Summary: Statistics for all residue classes")
    for filename, title in INPUT_FILES:
        sheet_name = title.replace(' ', '_').replace('≡', 'eq').replace('(', '').replace(')', '')
        print(f"  - {sheet_name}: Divisor table for {title}")

if __name__ == '__main__':
    main()