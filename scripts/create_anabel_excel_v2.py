#!/usr/bin/env python3
"""
create_anabel_excel_v2.py

Создание Excel файла для Anabel с делителями в виде СТЕПЕНЕЙ (не значений).
Формат: 12 вместо 4096, 13 вместо 8192, и т.д.

Запуск:
    python create_anabel_excel_v2.py

Выход:
    anabel_divisor_tables_v2.xlsx
"""

import os
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import math

# ============================================================================
# CONFIGURATION
# ============================================================================

# Input files (ONLY files that actually exist!)
INPUT_FILES = [
    ('septembrino_filtered_k_eq_1_mod_32.txt', 'k_eq_1_mod_32'),
    ('septembrino_filtered_k_eq_17_mod_32.txt', 'k_eq_17_mod_32'),
    ('septembrino_filtered_k_eq_3_mod_16.txt', 'k_eq_3_mod_16'),
]

OUTPUT_FILE = 'anabel_divisor_tables_v2.xlsx'
MAX_M = 60
MIN_DIVISOR = 32

# ============================================================================
# Helper functions
# ============================================================================

def divisor_to_exponent(d):
    """Convert divisor to exponent (e.g., 4096 → 12, 8192 → 13)."""
    if d < 32:
        return '.'
    try:
        exp = int(math.log2(d))
        return exp
    except:
        return d

def get_color_for_exponent(exp):
    """Get cell color based on exponent value."""
    if exp == '.':
        return None
    elif exp <= 5:
        return '92D050'
    elif exp <= 6:
        return 'C6E0B4'
    elif exp <= 7:
        return 'FFEB9C'
    elif exp <= 8:
        return 'FFC000'
    elif exp <= 9:
        return 'FF9900'
    elif exp <= 10:
        return 'FF6600'
    elif exp <= 11:
        return 'FF3300'
    elif exp <= 12:
        return 'FF0000'
    elif exp <= 13:
        return 'CC0000'
    elif exp <= 14:
        return '990000'
    elif exp <= 15:
        return '660000'
    elif exp <= 16:
        return '330000'
    else:
        return '000000'

def sanitize_sheet_name(name):
    """Remove invalid characters from sheet name."""
    invalid_chars = ['\\', '/', '*', '?', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '')
    name = name.replace('(', '').replace(')', '').replace('≡', '_eq_').replace(' ', '_')
    # Excel limit: 31 characters
    return name[:31]

# ============================================================================
# Parse filtered files
# ============================================================================

def parse_filtered_file(filename):
    """
    Parse the vertical list format.
    Returns: dict[k] = {m: highest_divisor}
    """
    data = {}
    
    if not os.path.exists(filename):
        print(f"WARNING: {filename} not found!")
        return data
    
    k_m_counter = {}
    
    with open(filename, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            
            if not line:
                continue
            
            if line.startswith('DIVISORS') or line.startswith('=') or line.startswith('Range'):
                continue
            
            if ':' not in line:
                continue
            
            parts = line.split(':', 1)
            if len(parts) != 2:
                continue
            
            k_str = parts[0].strip()
            divisors_str = parts[1].strip()
            
            try:
                k = int(k_str)
            except ValueError:
                continue
            
            if k not in k_m_counter:
                k_m_counter[k] = 0
            m = k_m_counter[k]
            k_m_counter[k] += 1
            
            if divisors_str:
                divisors = divisors_str.split()
                highest = 0
                for d in divisors:
                    try:
                        divisor = int(d)
                        if divisor >= MIN_DIVISOR and divisor > highest:
                            highest = divisor
                    except ValueError:
                        continue
                
                if k not in data:
                    data[k] = {}
                if highest > 0:
                    data[k][m] = highest
    
    return data

# ============================================================================
# Create Excel workbook
# ============================================================================

def create_excel_workbook(all_data, output_file):
    """Create Excel workbook with exponent values instead of divisors."""
    
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    for filename, sheet_name in INPUT_FILES:
        data = all_data.get(filename, {})
        if data:
            safe_name = sanitize_sheet_name(sheet_name)
            ws = wb.create_sheet(safe_name)
            _create_sheet(ws, data, sheet_name)
    
    wb.save(output_file)
    print(f"Created {output_file}")

def _create_sheet(ws, data, title):
    """Create a single sheet with exponent values."""
    
    ws['A1'] = f'Divisor Exponents: {title}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Value = 2^exponent (e.g., 12 = 4096, 13 = 8192)'
    ws['A2'].font = Font(italic=True)
    
    # Column headers (m = 0 to MAX_M)
    ws.cell(row=4, column=1).value = 'k'
    for m in range(MAX_M + 1):
        col = m + 2  # Column B = 2, C = 3, ...
        ws.cell(row=4, column=col).value = m
        ws.cell(row=4, column=col).font = Font(bold=True, size=9)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=4, column=col).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    sorted_k = sorted(data.keys())
    
    for row_idx, k in enumerate(sorted_k):
        excel_row = row_idx + 5
        k_data = data[k]
        
        ws.cell(row=excel_row, column=1).value = k
        ws.cell(row=excel_row, column=1).font = Font(bold=True)
        ws.cell(row=excel_row, column=1).alignment = Alignment(horizontal='right')
        
        for m in range(MAX_M + 1):
            excel_col = m + 2
            cell = ws.cell(row=excel_row, column=excel_col)
            
            if m in k_data:
                d = k_data[m]
                exp = divisor_to_exponent(d)
                cell.value = exp
                
                color = get_color_for_exponent(exp)
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                if exp != '.' and exp >= 12:
                    cell.font = Font(bold=True, color='FFFFFF', size=9)
                elif exp != '.':
                    cell.font = Font(bold=True, size=9)
                else:
                    cell.font = Font(size=9, color='999999')
                
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.value = '.'
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(size=9, color='999999')
        
        ws.row_dimensions[excel_row].height = 15
    
    # Set column widths using openpyxl's get_column_letter
    ws.column_dimensions['A'].width = 8
    for m in range(MAX_M + 1):
        col_letter = get_column_letter(m + 2)  # B=2, C=3, ..., BK=62
        ws.column_dimensions[col_letter].width = 5
    
    ws.freeze_panes = 'B5'

# ============================================================================
# Main
# ============================================================================

def main():
    print("=" * 80)
    print("CREATING EXCEL TABLES FOR ANABEL (v2 - EXPONENTS)")
    print("=" * 80)
    print()
    
    all_data = {}
    for filename, sheet_name in INPUT_FILES:
        print(f"Loading {filename}...")
        data = parse_filtered_file(filename)
        all_data[filename] = data
        print(f"  Loaded {len(data)} k values")
    print()
    
    print(f"Creating Excel workbook: {OUTPUT_FILE}...")
    create_excel_workbook(all_data, OUTPUT_FILE)
    print()
    
    print("=" * 80)
    print("DONE!")
    print("=" * 80)
    print()
    print(f"Open {OUTPUT_FILE} in Excel to view formatted tables.")
    print()
    print("Sheets:")
    for filename, sheet_name in INPUT_FILES:
        safe_name = sanitize_sheet_name(sheet_name)
        print(f"  - {safe_name}: {filename}")

if __name__ == '__main__':
    main()